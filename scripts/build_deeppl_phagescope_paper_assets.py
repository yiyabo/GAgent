#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import csv
import json
import math
import re
import shutil
import shlex
import tarfile
import textwrap
import time
import hashlib
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import matplotlib.pyplot as plt
import openpyxl
import requests
from matplotlib.colors import LinearSegmentedColormap
from dotenv import load_dotenv

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))
load_dotenv(PROJECT_ROOT / ".env")

TEST_LABELS_CSV = PROJECT_ROOT / "data" / "experiment_2" / "test_labels.csv"
SUPPLEMENTAL_XLSX = PROJECT_ROOT / "data" / "experiment_2" / "journal.pcbi.1012525.s001.xlsx"
DEEPPL_CODE_DIR = PROJECT_ROOT / "data" / "experiment_2" / "DeepPL"
DEFAULT_DEEPPL_MODEL_DIR = PROJECT_ROOT / "data" / "experiment_2" / "model" / "deeppl_ckpt-340000"
DEFAULT_OUTPUT_ROOT = PROJECT_ROOT / "docs" / "paper_assets" / "deeppl_phagescope_realrun_20260306"
EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
USER_AGENT = "GAgent-DeepPL-PhageScope-PaperAssets/1.0"
PAPER_METRICS = {
    "accuracy": 94.65,
    "sensitivity": 92.24,
    "specificity": 95.91,
    "f1": 0.92,
    "mcc": 0.53,
}
RAW_DEEPPL_FIELDNAMES = ["FASTA File", "Probability", "Prediction"]
MIN_ACCEPTED_DEEPPL_PREDICTIONS = 370
SUCCESS_STATES = {"SUCCESS", "SUCCEEDED", "COMPLETED", "DONE", "FINISHED"}
FAILED_STATES = {"FAILED", "ERROR"}
LYSOGENY_KEYWORDS = {
    "integrase": ["integrase"],
    "repressor": ["repressor"],
    "excisionase": ["excisionase", "xis"],
    "lysogeny": ["lysogen", "lysogenic", "prophage", "temperate"],
}
DETERMINISTIC_BASES = ("A", "C", "T", "G")
IUPAC_TO_BASES = {
    "R": ("A", "G"),
    "Y": ("C", "T"),
    "S": ("C", "G"),
    "W": ("A", "T"),
    "K": ("G", "T"),
    "M": ("A", "C"),
    "B": ("C", "G", "T"),
    "D": ("A", "G", "T"),
    "H": ("A", "C", "T"),
    "V": ("A", "C", "G"),
}
FIGURE_BG = "#f6f2ea"
PANEL_BG = "#fffdf8"
TEXT_PRIMARY = "#1f2937"
TEXT_MUTED = "#6b7280"
GRID_COLOR = "#d6d3cb"
BLUE_DARK = "#174b73"
BLUE_LIGHT = "#7db7d5"
GREEN = "#2b7a50"
RED = "#b84d3b"
GOLD = "#c9982e"
SLATE = "#94a3b8"


@dataclass
class Paths:
    root: Path
    ground_truth_dir: Path
    data_dir: Path
    test_set_fasta_dir: Path
    deeppl_dir: Path
    deeppl_raw_dir: Path
    deeppl_input_dir: Path
    phagescope_dir: Path
    phagescope_saveall_dir: Path
    integration_dir: Path
    figures_dir: Path
    methods_dir: Path
    manifest_path: Path


class PipelineError(RuntimeError):
    pass


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def build_paths(output_root: Path) -> Paths:
    return Paths(
        root=output_root,
        ground_truth_dir=output_root / "ground_truth",
        data_dir=output_root / "data",
        test_set_fasta_dir=output_root / "data" / "test_set_fasta",
        deeppl_dir=output_root / "deeppl",
        deeppl_raw_dir=output_root / "deeppl" / "raw",
        deeppl_input_dir=output_root / "deeppl" / "input_fasta",
        phagescope_dir=output_root / "phagescope",
        phagescope_saveall_dir=output_root / "phagescope" / "save_all",
        integration_dir=output_root / "integration",
        figures_dir=output_root / "figures",
        methods_dir=output_root / "methods",
        manifest_path=output_root / "manifest.json",
    )


def init_layout(paths: Paths) -> None:
    for path in (
        paths.root,
        paths.ground_truth_dir,
        paths.data_dir,
        paths.test_set_fasta_dir,
        paths.deeppl_dir,
        paths.deeppl_raw_dir,
        paths.deeppl_input_dir,
        paths.phagescope_dir,
        paths.phagescope_saveall_dir,
        paths.integration_dir,
        paths.figures_dir,
        paths.methods_dir,
    ):
        ensure_dir(path)


def load_manifest(path: Path) -> Dict[str, Any]:
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return {
        "generated_at": now_iso(),
        "updated_at": now_iso(),
        "historical_demo_only": [
            str((PROJECT_ROOT / "data" / "experiment_2" / "generate_replication_data.py").resolve()),
            str((PROJECT_ROOT / "docs" / "experiments_replicated" / "experiment_2_deeppl").resolve()),
        ],
        "inputs": {
            "test_labels_csv": str(TEST_LABELS_CSV.resolve()),
            "supplemental_xlsx": str(SUPPLEMENTAL_XLSX.resolve()),
        },
        "stages": {},
        "counts": {},
        "paths": {},
        "failures": [],
    }


def save_manifest(path: Path, manifest: Dict[str, Any]) -> None:
    manifest["updated_at"] = now_iso()
    path.write_text(json.dumps(manifest, indent=2, ensure_ascii=False), encoding="utf-8")


def mark_stage(manifest: Dict[str, Any], stage: str, status: str, **extra: Any) -> None:
    entry = dict(manifest.get("stages", {}).get(stage) or {})
    if status != "failed" and "error" not in extra:
        entry.pop("error", None)
    entry.update({"status": status, "updated_at": now_iso()})
    entry.update(extra)
    manifest.setdefault("stages", {})[stage] = entry


def append_failure(manifest: Dict[str, Any], stage: str, message: str) -> None:
    manifest.setdefault("failures", []).append({
        "stage": stage,
        "message": message,
        "time": now_iso(),
    })


def write_tsv(path: Path, fieldnames: List[str], rows: Iterable[Dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, delimiter="\t")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def read_tsv(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh, delimiter="\t"))


def read_csv_rows(path: Path) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as fh:
        return list(csv.DictReader(fh))


def write_csv_rows(path: Path, fieldnames: List[str], rows: Iterable[Dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def normalize_lifecycle_label(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        raise PipelineError("Empty lifecycle label")
    if any(token in text for token in ("lysogenic", "temperate", "lysogen")):
        return "temperate"
    if any(token in text for token in ("lytic", "virulent")):
        return "virulent"
    raise PipelineError(f"Unsupported lifecycle label: {value}")


def load_truth_from_csv() -> Dict[str, Dict[str, str]]:
    rows: Dict[str, Dict[str, str]] = {}
    with TEST_LABELS_CSV.open("r", encoding="utf-8", newline="") as fh:
        for row in csv.DictReader(fh):
            accession = str(row.get("accession") or "").strip()
            if not accession:
                continue
            rows[accession] = {
                "accession": accession,
                "lifecycle_raw": str(row.get("lifecycle") or "").strip(),
                "lifecycle_normalized": normalize_lifecycle_label(row.get("lifecycle")),
                "usage": str(row.get("usage") or "").strip(),
                "source": "test_labels.csv",
            }
    return rows


def load_truth_from_xlsx() -> Dict[str, Dict[str, str]]:
    workbook = openpyxl.load_workbook(SUPPLEMENTAL_XLSX, read_only=True, data_only=True)
    if "Test dataset" not in workbook.sheetnames:
        raise PipelineError("Supplemental workbook missing 'Test dataset' sheet")
    sheet = workbook["Test dataset"]
    header: Optional[List[str]] = None
    rows: Dict[str, Dict[str, str]] = {}
    for record in sheet.iter_rows(values_only=True):
        values = list(record)
        if header is None:
            header = [str(item or "").strip() for item in values]
            continue
        mapped = {header[idx]: values[idx] for idx in range(min(len(header), len(values)))}
        accession = str(mapped.get("Accession number") or "").strip()
        if not accession:
            continue
        rows[accession] = {
            "accession": accession,
            "lifecycle_raw": str(mapped.get("Lifecycle") or "").strip(),
            "lifecycle_normalized": normalize_lifecycle_label(mapped.get("Lifecycle")),
            "usage": str(mapped.get("Usage") or "").strip(),
            "source": "journal.pcbi.1012525.s001.xlsx",
        }
    return rows


def build_ground_truth(paths: Paths, manifest: Dict[str, Any]) -> List[Dict[str, str]]:
    csv_truth = load_truth_from_csv()
    xlsx_truth = load_truth_from_xlsx()
    if set(csv_truth) != set(xlsx_truth):
        missing_csv = sorted(set(xlsx_truth) - set(csv_truth))
        missing_xlsx = sorted(set(csv_truth) - set(xlsx_truth))
        raise PipelineError(
            "Ground-truth accession mismatch between CSV and XLSX: "
            f"missing_in_csv={missing_csv[:10]}, missing_in_xlsx={missing_xlsx[:10]}"
        )

    merged: List[Dict[str, str]] = []
    for accession in sorted(csv_truth):
        csv_row = csv_truth[accession]
        xlsx_row = xlsx_truth[accession]
        if csv_row["lifecycle_normalized"] != xlsx_row["lifecycle_normalized"]:
            raise PipelineError(
                f"Lifecycle mismatch for {accession}: {csv_row['lifecycle_raw']} vs {xlsx_row['lifecycle_raw']}"
            )
        merged.append(
            {
                "accession": accession,
                "lifecycle_raw": csv_row["lifecycle_raw"],
                "lifecycle_normalized": csv_row["lifecycle_normalized"],
                "usage": csv_row["usage"] or xlsx_row["usage"],
                "source": "test_labels.csv;journal.pcbi.1012525.s001.xlsx",
            }
        )

    out_path = paths.ground_truth_dir / "test_set.tsv"
    write_tsv(
        out_path,
        ["accession", "lifecycle_raw", "lifecycle_normalized", "usage", "source"],
        merged,
    )
    manifest.setdefault("paths", {})["ground_truth"] = str(out_path.resolve())
    manifest.setdefault("counts", {})["ground_truth_rows"] = len(merged)
    mark_stage(manifest, "ground_truth", "completed", count=len(merged), output=str(out_path.resolve()))
    return merged


def extract_first_fasta_record(text: str) -> Tuple[str, str]:
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines or not lines[0].startswith(">"):
        raise PipelineError("Downloaded payload is not FASTA")
    header = lines[0][1:].strip() or "unknown"
    seq_parts: List[str] = []
    for line in lines[1:]:
        if line.startswith(">"):
            break
        seq_parts.append(line)
    sequence = re.sub(r"\s+", "", "".join(seq_parts)).upper()
    if not sequence:
        raise PipelineError("Downloaded FASTA record has empty sequence")
    return header, sequence


def _choose_replacement_base(char: str, accession: str, index: int) -> str:
    candidates = IUPAC_TO_BASES.get(char, DETERMINISTIC_BASES)
    token = f"{accession}:{index}:{char}".encode("utf-8")
    digest = hashlib.sha256(token).hexdigest()
    return candidates[int(digest[:8], 16) % len(candidates)]


def prepare_deeppl_sequence(sequence: str, *, accession: str) -> Tuple[str, int, int, int]:
    compact = re.sub(r"\s+", "", sequence).upper()
    cleaned_chars: List[str] = []
    removed_n = 0
    replaced_ambiguous = 0
    unsupported: set[str] = set()
    for index, char in enumerate(compact):
        if char in {"A", "C", "T", "G"}:
            cleaned_chars.append(char)
        elif char == "N":
            removed_n += 1
        elif char in IUPAC_TO_BASES:
            replaced_ambiguous += 1
            if replaced_ambiguous > 10:
                raise PipelineError(f"Too many ambiguous bases for DeepPL input: {replaced_ambiguous}")
            cleaned_chars.append(_choose_replacement_base(char, accession, index))
        else:
            unsupported.add(char)
    if unsupported:
        raise PipelineError(f"Unsupported bases for DeepPL input: {''.join(sorted(unsupported))}")
    cleaned = "".join(cleaned_chars)
    if len(cleaned) < 106:
        raise PipelineError(f"Sequence too short for DeepPL after removing N: {len(cleaned)}")
    return cleaned, len(compact), removed_n, replaced_ambiguous


def download_test_set(paths: Paths, manifest: Dict[str, Any], truth_rows: List[Dict[str, str]]) -> List[Dict[str, Any]]:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    download_rows: List[Dict[str, Any]] = []
    for index, row in enumerate(truth_rows, start=1):
        accession = row["accession"]
        output_fasta = paths.test_set_fasta_dir / f"{accession}.fasta"
        deeppl_input = paths.deeppl_input_dir / f"{accession}.fasta"
        status = "ok"
        error = ""
        sequence_length = 0
        removed_n_count = 0
        replaced_ambiguous_count = 0
        source_header = ""
        try:
            if output_fasta.exists():
                header, sequence = extract_first_fasta_record(output_fasta.read_text(encoding="utf-8"))
                source_header = header
                sequence_length = len(sequence)
            else:
                response = session.get(
                    EFETCH_URL,
                    params={
                        "db": "nuccore",
                        "id": accession,
                        "rettype": "fasta",
                        "retmode": "text",
                    },
                    timeout=60,
                )
                response.raise_for_status()
                header, sequence = extract_first_fasta_record(response.text)
                source_header = header
                sequence_length = len(sequence)
                output_fasta.write_text(f">{accession}\n{sequence}\n", encoding="utf-8")
                time.sleep(0.34)
            cleaned, _, removed_n_count, replaced_ambiguous_count = prepare_deeppl_sequence(sequence, accession=accession)
            deeppl_input.write_text(f">{accession}\n{cleaned}\n", encoding="utf-8")
        except Exception as exc:  # pragma: no cover - network + remote inputs
            status = "failed"
            error = str(exc)
        download_rows.append(
            {
                "accession": accession,
                "lifecycle_normalized": row["lifecycle_normalized"],
                "status": status,
                "sequence_length": sequence_length,
                "removed_n_count": removed_n_count,
                "replaced_ambiguous_count": replaced_ambiguous_count,
                "source_header": source_header,
                "raw_fasta": str(output_fasta.resolve()) if output_fasta.exists() else "",
                "deeppl_input_fasta": str(deeppl_input.resolve()) if deeppl_input.exists() else "",
                "error": error,
            }
        )
        if index % 25 == 0:
            print(f"[download] {index}/{len(truth_rows)} processed")

    out_path = paths.data_dir / "download_manifest.tsv"
    write_tsv(
        out_path,
        [
            "accession",
            "lifecycle_normalized",
            "status",
            "sequence_length",
            "removed_n_count",
            "replaced_ambiguous_count",
            "source_header",
            "raw_fasta",
            "deeppl_input_fasta",
            "error",
        ],
        download_rows,
    )
    success_count = sum(1 for row in download_rows if row["status"] == "ok")
    failed_rows = [row for row in download_rows if row["status"] != "ok"]
    skipped_path = paths.data_dir / "download_skipped.tsv"
    write_tsv(
        skipped_path,
        [
            "accession",
            "lifecycle_normalized",
            "status",
            "sequence_length",
            "removed_n_count",
            "replaced_ambiguous_count",
            "source_header",
            "raw_fasta",
            "deeppl_input_fasta",
            "error",
        ],
        failed_rows,
    )
    manifest.setdefault("paths", {})["download_manifest"] = str(out_path.resolve())
    manifest.setdefault("paths", {})["download_skipped"] = str(skipped_path.resolve())
    manifest.setdefault("counts", {})["download_success"] = success_count
    manifest.setdefault("counts", {})["download_failed"] = len(failed_rows)
    mark_stage(
        manifest,
        "download_test_set",
        "completed" if success_count == len(truth_rows) else "partial",
        count=success_count,
        failed=len(failed_rows),
        output=str(out_path.resolve()),
    )
    return download_rows


def _deeppl_pending_run_path(paths: Paths) -> Path:
    return paths.deeppl_raw_dir / "pending_run.json"


async def _query_remote_deeppl_batch_status(
    config: Any,
    auth: Any,
    *,
    remote_output_csv: str,
    remote_log_path: str,
    pid: Optional[int],
) -> Dict[str, Any]:
    from tool_box.bio_tools.remote_executor import execute_remote_command

    pid_fragment = str(pid) if pid is not None else ""
    command = textwrap.dedent(
        f"""
        count=0
        if [ -f {shlex.quote(remote_output_csv)} ]; then
          count=$(wc -l < {shlex.quote(remote_output_csv)})
          if [ "$count" -gt 0 ]; then
            count=$((count - 1))
          fi
        fi
        running=0
        if [ -n {shlex.quote(pid_fragment)} ] && ps -p {shlex.quote(pid_fragment)} >/dev/null 2>&1; then
          running=1
        fi
        echo "__COUNT__=$count"
        echo "__RUNNING__=$running"
        if [ -f {shlex.quote(remote_log_path)} ]; then
          echo "__LOG_BEGIN__"
          tail -n 40 {shlex.quote(remote_log_path)}
          echo "__LOG_END__"
        fi
        """
    ).strip()
    result = await execute_remote_command(config, auth, command, timeout=60)
    stdout = str(result.get("stdout") or "")
    count_match = re.search(r"__COUNT__=(\d+)", stdout)
    running_match = re.search(r"__RUNNING__=(\d+)", stdout)
    log_match = re.search(r"__LOG_BEGIN__\n(?P<log>.*)\n__LOG_END__", stdout, flags=re.DOTALL)
    return {
        "success": bool(result.get("success")),
        "count": int(count_match.group(1)) if count_match else 0,
        "running": bool(int(running_match.group(1))) if running_match else False,
        "log_tail": (log_match.group("log") if log_match else "").strip(),
        "stderr": str(result.get("stderr") or ""),
    }


async def run_remote_deeppl_batch(paths: Paths, manifest: Dict[str, Any], *, remote_profile: str, timeout_sec: int) -> Dict[str, Any]:
    from tool_box.bio_tools.remote_executor import (
        RemoteExecutionConfig,
        create_remote_run_dirs,
        download_remote_run_dir,
        execute_remote_command,
        resolve_auth,
        upload_files,
    )
    from tool_box.tools_impl.deeppl import (
        _build_remote_env_keys,
        _normalize_remote_profile,
        _resolve_remote_int,
        _resolve_remote_string,
    )

    profile = _normalize_remote_profile(remote_profile)
    base = RemoteExecutionConfig.from_env()
    host = _resolve_remote_string(None, *_build_remote_env_keys(profile, "HOST", include_bio_fallback=True), fallback=base.host)
    user = _resolve_remote_string(None, *_build_remote_env_keys(profile, "USER", include_bio_fallback=True), fallback=base.user)
    port = _resolve_remote_int(None, *_build_remote_env_keys(profile, "PORT", include_bio_fallback=True), fallback=base.port)
    runtime_dir = _resolve_remote_string(None, *_build_remote_env_keys(profile, "RUNTIME_DIR", include_bio_fallback=True), fallback=base.runtime_dir.rstrip("/") + "/deeppl_batch")
    local_artifact_root = _resolve_remote_string(None, *_build_remote_env_keys(profile, "LOCAL_ARTIFACT_ROOT", include_bio_fallback=True), fallback=base.local_artifact_root)
    ssh_key_path = _resolve_remote_string(None, *_build_remote_env_keys(profile, "SSH_KEY_PATH", include_bio_fallback=True), fallback=base.ssh_key_path or "")
    password = _resolve_remote_string(None, *_build_remote_env_keys(profile, "PASSWORD", include_bio_fallback=True), fallback=base.password or "")
    remote_project = _resolve_remote_string(None, *_build_remote_env_keys(profile, "PROJECT_DIR"), fallback=str(DEEPPL_CODE_DIR))
    remote_script = _resolve_remote_string(None, *_build_remote_env_keys(profile, "BATCH_SCRIPT"), fallback=remote_project.rstrip("/") + "/mutilpredict_lyso_vs_lytic.py")
    remote_model_path = _resolve_remote_string(None, *_build_remote_env_keys(profile, "MODEL_PATH"), "DEEPPL_REMOTE_MODEL_PATH", fallback=str(DEFAULT_DEEPPL_MODEL_DIR))
    remote_python = _resolve_remote_string(None, *_build_remote_env_keys(profile, "PYTHON_BIN"), fallback="python")
    remote_pythonpath = _resolve_remote_string(None, *_build_remote_env_keys(profile, "PYTHONPATH"), fallback="")
    python_bin_dir = ""
    if "/" in remote_python:
        python_bin_dir = remote_python.rsplit("/", 1)[0].strip()
    path_export_prefix = f"export PATH={shlex.quote(python_bin_dir)}:$PATH && " if python_bin_dir else ""

    config = RemoteExecutionConfig(
        host=host,
        user=user,
        port=port,
        runtime_dir=runtime_dir,
        local_artifact_root=local_artifact_root,
        ssh_key_path=ssh_key_path or None,
        password=password or None,
        sudo_policy="never",
        connect_timeout=base.connect_timeout,
        scp_retries=base.scp_retries,
        scp_retry_delay=base.scp_retry_delay,
    )
    missing = config.missing_required()
    if missing:
        raise PipelineError("Missing remote DeepPL configuration: " + ", ".join(missing))

    auth = await resolve_auth(config)
    expected_rows = len(list(paths.deeppl_input_dir.glob("*.fasta")))
    pending_path = _deeppl_pending_run_path(paths)
    pending: Dict[str, Any] = {}

    if pending_path.exists():
        pending = json.loads(pending_path.read_text(encoding="utf-8"))
    else:
        run_id = "paper_assets_" + datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        remote_run_dir = f"{runtime_dir.rstrip('/')}/{run_id}"
        create_result = await create_remote_run_dirs(config, auth, remote_run_dir)
        if not create_result.get("success"):
            raise PipelineError("Failed to create remote DeepPL run dir: " + str(create_result.get("stderr") or create_result.get("error") or "unknown error"))

        archive_path = paths.deeppl_raw_dir / "deeppl_input_fasta.tar.gz"
        with tarfile.open(archive_path, "w:gz") as tar:
            for fasta in sorted(paths.deeppl_input_dir.glob("*.fasta")):
                tar.add(fasta, arcname=fasta.name)

        remote_archive = f"{remote_run_dir}/input/deeppl_input_fasta.tar.gz"
        upload_results = await upload_files(config, auth, [(str(archive_path), remote_archive)])
        if not upload_results or not upload_results[0].get("success"):
            failed = upload_results[0] if upload_results else {}
            raise PipelineError("Failed to upload DeepPL input archive: " + str(failed.get("stderr") or failed.get("error") or "unknown error"))

        remote_fasta_dir = f"{remote_run_dir}/input/fasta"
        remote_output_csv = f"{remote_run_dir}/output/deeppl_batch_predictions.csv"
        remote_log_path = f"{remote_run_dir}/output/deeppl_batch.log"
        export_pythonpath = f"export PYTHONPATH={shlex.quote(remote_pythonpath)}:$PYTHONPATH && " if remote_pythonpath else ""
        batch_command = (
            f"mkdir -p {shlex.quote(remote_fasta_dir)} && "
            f"tar --warning=no-timestamp -xzf {shlex.quote(remote_archive)} -C {shlex.quote(remote_fasta_dir)} && "
            f"cd {shlex.quote(remote_project)} && "
            f"{path_export_prefix}{export_pythonpath}{shlex.quote(remote_python)} {shlex.quote(remote_script)} "
            f"--model_path {shlex.quote(remote_model_path)} "
            f"--fasta_folder {shlex.quote(remote_fasta_dir)} "
            f"--output_csv {shlex.quote(remote_output_csv)}"
        )
        start_command = (
            f"nohup bash -lc {shlex.quote(batch_command)} "
            f"> {shlex.quote(remote_log_path)} 2>&1 < /dev/null & echo $!"
        )
        start_result = await execute_remote_command(config, auth, start_command, timeout=60)
        if not start_result.get("success"):
            raise PipelineError("Failed to start detached DeepPL batch: " + str(start_result.get("stderr") or start_result.get("error") or "unknown error"))
        pid_text = str(start_result.get("stdout") or "").strip().splitlines()[-1].strip()
        if not pid_text.isdigit():
            raise PipelineError(f"Detached DeepPL batch did not return a PID: {pid_text!r}")
        pending = {
            "run_id": run_id,
            "time": now_iso(),
            "remote_profile": profile,
            "remote_host": host,
            "remote_user": user,
            "remote_port": port,
            "remote_project": remote_project,
            "remote_script": remote_script,
            "remote_model_path": remote_model_path,
            "remote_run_dir": remote_run_dir,
            "remote_output_csv": remote_output_csv,
            "remote_log_path": remote_log_path,
            "pid": int(pid_text),
            "expected_rows": expected_rows,
        }
        pending_path.write_text(json.dumps(pending, indent=2, ensure_ascii=False), encoding="utf-8")

    remote_run_dir = str(pending["remote_run_dir"])
    remote_output_csv = str(pending["remote_output_csv"])
    remote_log_path = str(pending["remote_log_path"])
    pid = int(pending["pid"]) if str(pending.get("pid") or "").isdigit() else None
    run_id = str(pending["run_id"])

    start = time.monotonic()
    last_count = -1
    current_count = 0
    last_log_tail = ""
    exited_early = False
    while True:
        status = await _query_remote_deeppl_batch_status(
            config,
            auth,
            remote_output_csv=remote_output_csv,
            remote_log_path=remote_log_path,
            pid=pid,
        )
        if not status["success"]:
            raise PipelineError("Failed to poll DeepPL batch status: " + status["stderr"])
        current_count = int(status["count"])
        last_log_tail = str(status.get("log_tail") or "")
        if current_count != last_count:
            print(f"[deeppl] progress {current_count}/{expected_rows}")
            last_count = current_count
        if current_count >= expected_rows:
            break
        if not status["running"]:
            exited_early = True
            break
        if time.monotonic() - start > timeout_sec:
            raise PipelineError(f"DeepPL remote batch timed out after {timeout_sec}s at {current_count}/{expected_rows}")
        await asyncio.sleep(30)

    local_remote_artifacts = paths.deeppl_raw_dir / "remote_artifacts" / run_id
    download_result = await download_remote_run_dir(config, auth, remote_run_dir, str(local_remote_artifacts))
    if not download_result.get("success"):
        raise PipelineError("Failed to download DeepPL remote artifacts: " + str(download_result.get("stderr") or download_result.get("error") or "unknown error"))

    downloaded_csv = local_remote_artifacts / "output" / "deeppl_batch_predictions.csv"
    if not downloaded_csv.exists():
        raise PipelineError(f"DeepPL batch output missing: {downloaded_csv}")
    local_csv = paths.deeppl_raw_dir / "deeppl_batch_predictions.csv"
    shutil.copy2(downloaded_csv, local_csv)
    if pending_path.exists():
        pending_path.unlink()

    provenance = {
        "run_id": run_id,
        "time": now_iso(),
        "remote_profile": profile,
        "remote_host": host,
        "remote_user": user,
        "remote_port": port,
        "remote_project": remote_project,
        "remote_script": remote_script,
        "remote_model_path": remote_model_path,
        "remote_run_dir": remote_run_dir,
        "local_artifact_dir": str(local_remote_artifacts.resolve()),
        "pid": pid,
        "completed_count": current_count,
        "expected_rows": expected_rows,
        "incomplete_exit": exited_early,
        "log_tail": last_log_tail,
        "stdout_preview": "",
        "stderr_preview": "",
    }
    run_provenance_path = paths.deeppl_dir / "run_provenance.json"
    run_provenance_path.write_text(json.dumps(provenance, indent=2, ensure_ascii=False), encoding="utf-8")
    manifest.setdefault("paths", {})["deeppl_run_provenance"] = str(run_provenance_path.resolve())
    return provenance


def compute_metrics(prediction_rows: List[Dict[str, Any]]) -> Dict[str, float]:
    tp = sum(1 for row in prediction_rows if row["true_label"] == "temperate" and row["deeppl_label"] == "temperate")
    tn = sum(1 for row in prediction_rows if row["true_label"] == "virulent" and row["deeppl_label"] == "virulent")
    fp = sum(1 for row in prediction_rows if row["true_label"] == "virulent" and row["deeppl_label"] == "temperate")
    fn = sum(1 for row in prediction_rows if row["true_label"] == "temperate" and row["deeppl_label"] == "virulent")
    n_total = len(prediction_rows)
    accuracy = (tp + tn) / n_total if n_total else 0.0
    sensitivity = tp / (tp + fn) if (tp + fn) else 0.0
    specificity = tn / (tn + fp) if (tn + fp) else 0.0
    precision = tp / (tp + fp) if (tp + fp) else 0.0
    f1 = 2 * precision * sensitivity / (precision + sensitivity) if (precision + sensitivity) else 0.0
    denom = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = ((tp * tn) - (fp * fn)) / denom if denom else 0.0
    return {
        "n_total": n_total,
        "tp": tp,
        "tn": tn,
        "fp": fp,
        "fn": fn,
        "accuracy": round(accuracy * 100, 4),
        "sensitivity": round(sensitivity * 100, 4),
        "specificity": round(specificity * 100, 4),
        "precision": round(precision * 100, 4),
        "f1": round(f1, 6),
        "mcc": round(mcc, 6),
    }


def parse_deeppl_raw_rows(
    raw_rows: List[Dict[str, Any]],
    truth_by_accession: Dict[str, Dict[str, str]],
) -> Tuple[List[Dict[str, Any]], List[str]]:
    by_accession: Dict[str, Dict[str, Any]] = {}
    for raw in raw_rows:
        fasta_file = str(raw.get("FASTA File") or "").strip()
        accession = Path(fasta_file).stem
        if not accession:
            continue
        if accession not in truth_by_accession:
            raise PipelineError(f"Unexpected DeepPL prediction accession: {accession}")
        prediction_text = str(raw.get("Prediction") or "").strip()
        probability = float(raw.get("Probability") or 0.0)
        by_accession[accession] = {
            "accession": accession,
            "true_label": truth_by_accession[accession]["lifecycle_normalized"],
            "deeppl_raw_label": prediction_text,
            "deeppl_label": normalize_lifecycle_label(prediction_text),
            "positive_window_fraction": probability,
            "window_score_threshold": 0.9,
            "positive_window_fraction_threshold": 0.016,
            "prediction_source": "mutilpredict_lyso_vs_lytic.py",
        }

    benchmark_rows = sorted(by_accession.values(), key=lambda item: item["accession"])
    missing = sorted(set(truth_by_accession) - set(by_accession))
    return benchmark_rows, missing


def merge_deeppl_raw_rows(existing_rows: List[Dict[str, Any]], supplemental_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: Dict[str, Dict[str, Any]] = {}
    for row in existing_rows:
        fasta_file = str(row.get("FASTA File") or "").strip()
        if fasta_file:
            merged[fasta_file] = {
                "FASTA File": fasta_file,
                "Probability": row.get("Probability", ""),
                "Prediction": row.get("Prediction", ""),
            }
    for row in supplemental_rows:
        fasta_file = str(row.get("FASTA File") or "").strip()
        if fasta_file:
            merged[fasta_file] = {
                "FASTA File": fasta_file,
                "Probability": row.get("Probability", ""),
                "Prediction": row.get("Prediction", ""),
            }
    return [merged[key] for key in sorted(merged)]


async def supplement_missing_deeppl_predictions(
    paths: Paths,
    *,
    missing_accessions: List[str],
    remote_profile: str,
    timeout_sec: int,
    manifest: Dict[str, Any],
) -> Tuple[List[Dict[str, Any]], Dict[str, str]]:
    from tool_box.tools_impl.deeppl import deeppl_handler

    if not missing_accessions:
        return [], {}

    pending_meta: Dict[str, Any] = {}
    pending_path = _deeppl_pending_run_path(paths)
    if pending_path.exists():
        pending_meta = json.loads(pending_path.read_text(encoding="utf-8"))
    provenance_path = paths.deeppl_dir / "run_provenance.json"
    if provenance_path.exists():
        pending_meta.update(json.loads(provenance_path.read_text(encoding="utf-8")))

    supplemental_rows: List[Dict[str, Any]] = []
    failures: Dict[str, str] = {}
    per_item_timeout = max(900, min(timeout_sec, 5400))

    for accession in missing_accessions:
        input_fasta = paths.deeppl_input_dir / f"{accession}.fasta"
        if not input_fasta.exists():
            failures[accession] = f"Missing normalized input FASTA: {input_fasta}"
            continue

        result = await deeppl_handler(
            action="predict",
            input_file=str(input_fasta),
            execution_mode="remote",
            remote_profile=remote_profile,
            model_path=pending_meta.get("remote_model_path"),
            remote_project_dir=pending_meta.get("remote_project"),
            remote_python_bin=None,
            remote_password=None,
            timeout=per_item_timeout,
            sample_id=accession,
        )
        if not result.get("success"):
            failures[accession] = str(result.get("error") or result.get("stderr") or "DeepPL supplemental prediction failed")
            print(f"[deeppl] supplemental retry failed for {accession}: {failures[accession]}")
            continue

        predicted_label = str(result.get("predicted_label") or "").strip().lower()
        if predicted_label.startswith("lyso"):
            prediction_text = "Lysogenic"
        elif predicted_label.startswith("lyt"):
            prediction_text = "Lytic"
        else:
            failures[accession] = f"Unsupported supplemental label: {predicted_label}"
            print(f"[deeppl] supplemental retry failed for {accession}: {failures[accession]}")
            continue

        supplemental_rows.append(
            {
                "FASTA File": f"{accession}.fasta",
                "Probability": str(result.get("positive_window_fraction")),
                "Prediction": prediction_text,
            }
        )
        print(f"[deeppl] supplemental retry succeeded for {accession}")

    if failures:
        append_failure(
            manifest,
            "deeppl_benchmark",
            "Supplemental DeepPL predictions failed: "
            + "; ".join(f"{key}: {value}" for key, value in sorted(failures.items())),
        )
    return supplemental_rows, failures


def should_accept_partial_deeppl_benchmark(prediction_count: int) -> bool:
    return prediction_count >= MIN_ACCEPTED_DEEPPL_PREDICTIONS


def build_deeppl_benchmark(paths: Paths, manifest: Dict[str, Any], truth_rows: List[Dict[str, str]], download_rows: List[Dict[str, Any]], *, remote_profile: str, timeout_sec: int) -> List[Dict[str, Any]]:
    ok_inputs = [row for row in download_rows if row["status"] == "ok"]
    skipped_inputs = [row for row in download_rows if row["status"] != "ok"]
    if not ok_inputs:
        raise PipelineError("DeepPL input set is empty after download/preprocessing")

    raw_csv = paths.deeppl_raw_dir / "deeppl_batch_predictions.csv"
    benchmark_path = paths.deeppl_dir / "benchmark_predictions.tsv"
    metrics_path = paths.deeppl_dir / "benchmark_metrics.json"
    confusion_path = paths.deeppl_dir / "confusion_matrix.tsv"
    missing_path = paths.deeppl_dir / "missing_predictions.tsv"
    if raw_csv.exists() and benchmark_path.exists() and metrics_path.exists() and confusion_path.exists():
        benchmark_rows = read_tsv(benchmark_path)
        if len(benchmark_rows) == len(ok_inputs) or should_accept_partial_deeppl_benchmark(len(benchmark_rows)):
            manifest.setdefault("paths", {}).update(
                {
                    "deeppl_raw_csv": str(raw_csv.resolve()),
                    "deeppl_benchmark_predictions": str(benchmark_path.resolve()),
                    "deeppl_benchmark_metrics": str(metrics_path.resolve()),
                    "deeppl_confusion_matrix": str(confusion_path.resolve()),
                }
            )
            manifest.setdefault("counts", {})["deeppl_predictions"] = len(benchmark_rows)
            manifest.setdefault("counts", {})["deeppl_prediction_omitted"] = max(0, len(ok_inputs) - len(benchmark_rows))
            if missing_path.exists():
                manifest.setdefault("paths", {})["deeppl_missing_predictions"] = str(missing_path.resolve())
            mark_stage(
                manifest,
                "deeppl_benchmark",
                "completed" if len(benchmark_rows) == len(ok_inputs) else "partial",
                count=len(benchmark_rows),
                remote_profile=remote_profile,
                reused=True,
            )
            return [
                {
                    **row,
                    "positive_window_fraction": float(row["positive_window_fraction"]),
                    "window_score_threshold": float(row["window_score_threshold"]),
                    "positive_window_fraction_threshold": float(row["positive_window_fraction_threshold"]),
                }
                for row in benchmark_rows
            ]

    truth_by_accession = {row["accession"]: row for row in truth_rows if row["accession"] in {item["accession"] for item in ok_inputs}}
    if raw_csv.exists():
        provenance_path = paths.deeppl_dir / "run_provenance.json"
        provenance = json.loads(provenance_path.read_text(encoding="utf-8")) if provenance_path.exists() else {}
    else:
        provenance = asyncio.run(run_remote_deeppl_batch(paths, manifest, remote_profile=remote_profile, timeout_sec=timeout_sec))

    raw_rows = read_csv_rows(raw_csv)
    benchmark_rows, missing = parse_deeppl_raw_rows(raw_rows, truth_by_accession)

    if missing:
        supplemental_rows, _supplemental_failures = asyncio.run(
            supplement_missing_deeppl_predictions(
                paths,
                missing_accessions=missing,
                remote_profile=remote_profile,
                timeout_sec=timeout_sec,
                manifest=manifest,
            )
        )
        if supplemental_rows:
            raw_rows = merge_deeppl_raw_rows(raw_rows, supplemental_rows)
            write_csv_rows(raw_csv, RAW_DEEPPL_FIELDNAMES, raw_rows)
            benchmark_rows, missing = parse_deeppl_raw_rows(raw_rows, truth_by_accession)

    if missing:
        if not should_accept_partial_deeppl_benchmark(len(benchmark_rows)):
            raise PipelineError(
                f"DeepPL predictions missing {len(missing)} accessions after supplemental retries: {', '.join(missing[:10])}"
            )
        write_tsv(
            missing_path,
            ["accession", "true_label"],
            [{"accession": accession, "true_label": truth_by_accession[accession]["lifecycle_normalized"]} for accession in missing],
        )
        manifest.setdefault("paths", {})["deeppl_missing_predictions"] = str(missing_path.resolve())
        manifest.setdefault("counts", {})["deeppl_prediction_omitted"] = len(missing)
        append_failure(
            manifest,
            "deeppl_benchmark",
            "Proceeding with partial DeepPL benchmark; missing predictions: " + ", ".join(missing),
        )
    elif missing_path.exists():
        missing_path.unlink()

    benchmark_rows.sort(key=lambda item: item["accession"])
    write_tsv(
        benchmark_path,
        [
            "accession",
            "true_label",
            "deeppl_raw_label",
            "deeppl_label",
            "positive_window_fraction",
            "window_score_threshold",
            "positive_window_fraction_threshold",
            "prediction_source",
        ],
        benchmark_rows,
    )

    metrics = compute_metrics(benchmark_rows)
    metrics_path.write_text(json.dumps(metrics, indent=2, ensure_ascii=False), encoding="utf-8")

    write_tsv(
        confusion_path,
        ["true_label", "predicted_label", "count"],
        [
            {"true_label": "temperate", "predicted_label": "temperate", "count": metrics["tp"]},
            {"true_label": "temperate", "predicted_label": "virulent", "count": metrics["fn"]},
            {"true_label": "virulent", "predicted_label": "temperate", "count": metrics["fp"]},
            {"true_label": "virulent", "predicted_label": "virulent", "count": metrics["tn"]},
        ],
    )

    manifest.setdefault("paths", {}).update(
        {
            "deeppl_raw_csv": str(raw_csv.resolve()),
            "deeppl_benchmark_predictions": str(benchmark_path.resolve()),
            "deeppl_benchmark_metrics": str(metrics_path.resolve()),
            "deeppl_confusion_matrix": str(confusion_path.resolve()),
        }
    )
    manifest.setdefault("counts", {})["deeppl_predictions"] = len(benchmark_rows)
    manifest.setdefault("counts", {})["deeppl_skipped"] = len(skipped_inputs)
    if not missing:
        manifest.setdefault("counts", {})["deeppl_prediction_omitted"] = 0
    if skipped_inputs:
        append_failure(
            manifest,
            "deeppl_benchmark",
            "Skipped source records: " + ", ".join(f"{row['accession']} ({row['error']})" for row in skipped_inputs),
        )
    mark_stage(
        manifest,
        "deeppl_benchmark",
        "completed" if not skipped_inputs and not missing else "partial",
        count=len(benchmark_rows),
        skipped=len(skipped_inputs),
        omitted=len(missing),
        omitted_accessions=missing,
        remote_profile=remote_profile,
        provenance=provenance.get("run_id"),
    )
    return benchmark_rows


def select_validation_subset(benchmark_rows: List[Dict[str, Any]], per_class: int) -> List[Dict[str, Any]]:
    by_label: Dict[str, List[Dict[str, Any]]] = {"temperate": [], "virulent": []}
    for row in benchmark_rows:
        by_label[row["true_label"]].append(row)
    subset: List[Dict[str, Any]] = []
    for label in ("temperate", "virulent"):
        selected = sorted(by_label[label], key=lambda item: item["accession"])[:per_class]
        if len(selected) < per_class:
            raise PipelineError(f"Not enough samples for {label}: {len(selected)} < {per_class}")
        subset.extend(selected)
    subset.sort(key=lambda item: (item["true_label"], item["accession"]))
    return [
        {
            "sample_id": item["accession"],
            "accession": item["accession"],
            "true_label": item["true_label"],
            "deeppl_label": item["deeppl_label"],
            "positive_window_fraction": item["positive_window_fraction"],
            "input_fasta": str((DEFAULT_OUTPUT_ROOT / "deeppl" / "input_fasta" / f"{item['accession']}.fasta").resolve()),
        }
        for item in subset
    ]


# The input_fasta value above is rewritten to the actual output root before saving.

def normalize_subset_rows(paths: Paths, subset_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    normalized: List[Dict[str, Any]] = []
    for row in subset_rows:
        item = dict(row)
        item["input_fasta"] = str((paths.deeppl_input_dir / f"{row['accession']}.fasta").resolve())
        normalized.append(item)
    return normalized


async def submit_phagescope_sample(sample: Dict[str, Any]) -> Dict[str, Any]:
    from tool_box.tools_impl.phagescope import phagescope_handler

    return await phagescope_handler(
        action="submit",
        file_path=sample["input_fasta"],
        inputtype="upload",
        analysistype="Annotation Pipline",
        userid="paper_assets_pipeline",
        modulelist={"quality": True, "host": True, "lifestyle": True, "annotation": True},
        timeout=120.0,
    )


async def task_detail_phagescope(taskid: str) -> Dict[str, Any]:
    from tool_box.tools_impl.phagescope import phagescope_handler

    return await phagescope_handler(action="task_detail", taskid=taskid, timeout=60.0)


async def save_all_phagescope(taskid: str, save_path: Path) -> Dict[str, Any]:
    from tool_box.tools_impl.phagescope import phagescope_handler

    return await phagescope_handler(action="save_all", taskid=taskid, save_path=str(save_path), timeout=180.0)


def extract_taskid(result: Dict[str, Any]) -> Optional[str]:
    from app.routers.chat.session_helpers import _extract_taskid_from_result

    taskid = _extract_taskid_from_result(result)
    return str(taskid).strip() if taskid is not None else None


def extract_snapshot(detail_result: Dict[str, Any]) -> Dict[str, Any]:
    from app.routers.chat.session_helpers import _extract_phagescope_task_snapshot

    return _extract_phagescope_task_snapshot(detail_result)


def is_task_complete(snapshot: Dict[str, Any]) -> bool:
    counts = snapshot.get("counts") if isinstance(snapshot.get("counts"), dict) else {}
    total = counts.get("total") if isinstance(counts.get("total"), int) else 0
    done = counts.get("done") if isinstance(counts.get("done"), int) else 0
    failed = counts.get("failed") if isinstance(counts.get("failed"), int) else 0
    if total > 0 and (done + failed) >= total:
        return True
    status_text = str(snapshot.get("task_status") or snapshot.get("remote_status") or "").strip().upper()
    return status_text in SUCCESS_STATES


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def first_result_dict(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    payload = read_json(path)
    results = payload.get("results") if isinstance(payload, dict) else None
    if isinstance(results, list) and results and isinstance(results[0], dict):
        return results[0]
    return {}


def summarize_quality_payload(payload: Any) -> str:
    if not isinstance(payload, dict):
        return ""
    results = payload.get("results")
    if isinstance(results, dict):
        quality_summary = results.get("quality_summary")
        if isinstance(quality_summary, list) and quality_summary and isinstance(quality_summary[0], dict):
            first = quality_summary[0]
            return json.dumps(first, ensure_ascii=False, sort_keys=True)
    return ""


def extract_completed_sample(sample: Dict[str, Any], save_all_result: Dict[str, Any]) -> Dict[str, Any]:
    output_dir = Path(str(save_all_result.get("output_directory") or "")).resolve()
    phage_info_path = output_dir / "metadata" / "phage_info.json"
    quality_path = output_dir / "metadata" / "quality.json"
    if not quality_path.exists():
        quality_path = output_dir / "metadata" / "quality_from_modules.json"

    phagescope_lifestyle_raw = ""
    phagescope_lifestyle = ""
    host = ""
    row = first_result_dict(phage_info_path)
    if row:
        phagescope_lifestyle_raw = str(row.get("lifestyle") or "").strip()
        if phagescope_lifestyle_raw:
            try:
                phagescope_lifestyle = normalize_lifecycle_label(phagescope_lifestyle_raw)
            except PipelineError:
                phagescope_lifestyle = ""
        host = str(row.get("host") or "").strip()

    if not phagescope_lifestyle:
        for candidate in (
            output_dir / "annotation" / "module_lifestyle.json",
            output_dir / "raw_api_responses" / "modules_lifestyle_raw.json",
        ):
            row = first_result_dict(candidate)
            phagescope_lifestyle_raw = str(row.get("lifestyle") or "").strip()
            if phagescope_lifestyle_raw:
                try:
                    phagescope_lifestyle = normalize_lifecycle_label(phagescope_lifestyle_raw)
                except PipelineError:
                    phagescope_lifestyle = ""
                if phagescope_lifestyle:
                    break

    if not host:
        for candidate in (
            output_dir / "annotation" / "module_host.json",
            output_dir / "raw_api_responses" / "modules_host_raw.json",
        ):
            row = first_result_dict(candidate)
            host = str(row.get("host") or row.get("Species") or "").strip()
            if host:
                break

    quality_summary = ""
    if quality_path.exists():
        quality_summary = summarize_quality_payload(read_json(quality_path))

    return {
        "sample_id": sample["sample_id"],
        "accession": sample["accession"],
        "taskid": str(save_all_result.get("taskid") or sample.get("taskid") or ""),
        "status": "saved",
        "phagescope_lifestyle_raw": phagescope_lifestyle_raw,
        "phagescope_lifestyle": phagescope_lifestyle,
        "quality_summary": quality_summary,
        "host": host,
        "artifact_root": str(output_dir),
    }


def load_registry(path: Path) -> Dict[str, Dict[str, str]]:
    rows = read_tsv(path)
    return {row["sample_id"]: row for row in rows if row.get("sample_id")}


def save_registry(path: Path, rows: Dict[str, Dict[str, Any]]) -> None:
    fieldnames = [
        "sample_id",
        "accession",
        "true_label",
        "deeppl_label",
        "taskid",
        "status",
        "remote_status",
        "task_status",
        "done",
        "total",
        "submitted_at",
        "updated_at",
        "save_all_artifact_root",
        "last_error",
    ]
    ordered = [rows[key] for key in sorted(rows)]
    write_tsv(path, fieldnames, ordered)


def save_completed_samples(path: Path, rows: List[Dict[str, Any]]) -> None:
    write_tsv(
        path,
        [
            "sample_id",
            "accession",
            "taskid",
            "status",
            "phagescope_lifestyle_raw",
            "phagescope_lifestyle",
            "quality_summary",
            "host",
            "artifact_root",
        ],
        sorted(rows, key=lambda item: item["sample_id"]),
    )


def refresh_completed_samples(paths: Paths, registry: Dict[str, Dict[str, Any]], subset_rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    sample_by_id = {row["sample_id"]: row for row in subset_rows}
    refreshed: Dict[str, Dict[str, Any]] = {}
    for sample_id, entry in sorted(registry.items()):
        if str(entry.get("status") or "") != "saved":
            continue
        artifact_root = str(entry.get("save_all_artifact_root") or "").strip()
        if not artifact_root:
            continue
        sample = dict(sample_by_id.get(sample_id) or {})
        sample.setdefault("sample_id", sample_id)
        sample.setdefault("accession", str(entry.get("accession") or sample_id))
        sample.setdefault("taskid", str(entry.get("taskid") or ""))
        refreshed[sample_id] = extract_completed_sample(
            sample,
            {
                "taskid": str(entry.get("taskid") or ""),
                "output_directory": artifact_root,
            },
        )
    return refreshed


def build_subset_manifest(paths: Paths, benchmark_rows: List[Dict[str, Any]], per_class: int) -> List[Dict[str, Any]]:
    subset = normalize_subset_rows(paths, select_validation_subset(benchmark_rows, per_class))
    out_path = paths.phagescope_dir / "subset_manifest.tsv"
    write_tsv(
        out_path,
        ["sample_id", "accession", "true_label", "deeppl_label", "positive_window_fraction", "input_fasta"],
        subset,
    )
    return subset


async def run_phagescope_validation(paths: Paths, manifest: Dict[str, Any], subset_rows: List[Dict[str, Any]], *, poll_interval: float, poll_timeout: float) -> Tuple[List[Dict[str, Any]], Dict[str, Dict[str, Any]]]:
    registry_path = paths.phagescope_dir / "task_registry.tsv"
    completed_path = paths.phagescope_dir / "completed_samples.tsv"
    registry = load_registry(registry_path)
    completed_by_sample = {row["sample_id"]: row for row in read_tsv(completed_path)}
    if not completed_path.exists():
        save_completed_samples(completed_path, [])
    refreshed = refresh_completed_samples(paths, registry, subset_rows)
    if refreshed:
        completed_by_sample.update(refreshed)
        save_completed_samples(completed_path, list(completed_by_sample.values()))

    # Submit missing tasks first.
    for sample in subset_rows:
        existing = registry.get(sample["sample_id"])
        if existing and existing.get("taskid"):
            continue
        result = await submit_phagescope_sample(sample)
        taskid = extract_taskid(result) if result.get("success") else None
        registry[sample["sample_id"]] = {
            "sample_id": sample["sample_id"],
            "accession": sample["accession"],
            "true_label": sample["true_label"],
            "deeppl_label": sample["deeppl_label"],
            "taskid": taskid or "",
            "status": "submitted" if taskid else "submit_failed",
            "remote_status": "",
            "task_status": "",
            "done": "0",
            "total": "0",
            "submitted_at": now_iso(),
            "updated_at": now_iso(),
            "save_all_artifact_root": "",
            "last_error": "" if taskid else str(result.get("error") or "submit returned no taskid"),
        }
        save_registry(registry_path, registry)

    deadline = time.monotonic() + max(poll_timeout, 0.0)
    pending = {sample["sample_id"] for sample in subset_rows}
    while pending and time.monotonic() < deadline:
        progress_made = False
        for sample in subset_rows:
            sample_id = sample["sample_id"]
            if sample_id not in pending:
                continue
            entry = registry.get(sample_id) or {}
            taskid = str(entry.get("taskid") or "").strip()
            if not taskid:
                pending.remove(sample_id)
                continue
            if sample_id in completed_by_sample:
                pending.remove(sample_id)
                continue

            detail = await task_detail_phagescope(taskid)
            snapshot = extract_snapshot(detail)
            counts = snapshot.get("counts") if isinstance(snapshot.get("counts"), dict) else {}
            entry["remote_status"] = str(snapshot.get("remote_status") or "")
            entry["task_status"] = str(snapshot.get("task_status") or "")
            entry["done"] = str(counts.get("done") or 0)
            entry["total"] = str(counts.get("total") or 0)
            entry["updated_at"] = now_iso()
            save_registry(registry_path, registry)

            if not is_task_complete(snapshot):
                continue

            save_dir = paths.phagescope_saveall_dir / sample_id
            save_result = await save_all_phagescope(taskid, save_dir)
            if bool(save_result.get("success")):
                completed = extract_completed_sample(sample, save_result)
                completed_by_sample[sample_id] = completed
                entry["status"] = "saved"
                entry["save_all_artifact_root"] = completed["artifact_root"]
                entry["last_error"] = ""
                pending.remove(sample_id)
                progress_made = True
                save_completed_samples(completed_path, list(completed_by_sample.values()))
                save_registry(registry_path, registry)
            else:
                entry["status"] = "save_all_failed"
                entry["last_error"] = str(save_result.get("error") or "save_all failed")
                save_registry(registry_path, registry)

        if pending and not progress_made:
            await asyncio.sleep(max(poll_interval, 1.0))

    manifest.setdefault("paths", {})["phagescope_subset_manifest"] = str((paths.phagescope_dir / "subset_manifest.tsv").resolve())
    manifest.setdefault("paths", {})["phagescope_task_registry"] = str(registry_path.resolve())
    manifest.setdefault("paths", {})["phagescope_completed_samples"] = str(completed_path.resolve())
    mark_stage(
        manifest,
        "phagescope_validation",
        "completed" if not pending else "running",
        completed=len(completed_by_sample),
        pending=len(pending),
        total=len(subset_rows),
    )
    return list(completed_by_sample.values()), registry


def count_keyword_hits(root: Path) -> Dict[str, int]:
    content_parts: List[str] = []
    for rel in (
        Path("annotation/proteins.tsv"),
        Path("annotation/proteins.json"),
        Path("annotation/proteins_from_annotation.tsv"),
        Path("annotation/proteins_from_annotation.json"),
    ):
        path = root / rel
        if path.exists():
            content_parts.append(path.read_text(encoding="utf-8", errors="replace"))
    haystack = "\n".join(content_parts).lower()
    counts: Dict[str, int] = {}
    for key, variants in LYSOGENY_KEYWORDS.items():
        counts[key] = sum(haystack.count(variant) for variant in variants)
    return counts


def build_integration_outputs(paths: Paths, manifest: Dict[str, Any], benchmark_rows: List[Dict[str, Any]], completed_samples: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    benchmark_by_accession = {row["accession"]: row for row in benchmark_rows}
    comparison_rows: List[Dict[str, Any]] = []
    disagreement_rows: List[Dict[str, Any]] = []
    omitted_rows: List[Dict[str, Any]] = []

    for sample in sorted(completed_samples, key=lambda item: item["sample_id"]):
        accession = sample["accession"]
        if accession not in benchmark_by_accession:
            omitted_rows.append(
                {
                    "sample_id": sample["sample_id"],
                    "accession": accession,
                    "reason": "missing_benchmark_prediction",
                }
            )
            continue
        deeppl_row = benchmark_by_accession[accession]
        phagescope_label = sample.get("phagescope_lifestyle") or ""
        deeppl_label = deeppl_row.get("deeppl_label") or ""
        if not phagescope_label or not deeppl_label:
            omitted_rows.append(
                {
                    "sample_id": sample["sample_id"],
                    "accession": accession,
                    "reason": "missing_phagescope_lifestyle" if not phagescope_label else "missing_deeppl_label",
                }
            )
            continue
        consensus = "agree" if phagescope_label == deeppl_label else "disagree"
        confidence = "high" if consensus == "agree" else "needs_review"
        notes = (
            "DeepPL and PhageScope lifecycle labels agree."
            if consensus == "agree"
            else "DeepPL and PhageScope lifecycle labels disagree; review integrase/repressor evidence."
        )
        comparison_rows.append(
            {
                "sample_id": sample["sample_id"],
                "accession": accession,
                "phagescope_lifestyle": phagescope_label,
                "deeppl_label": deeppl_label,
                "consensus": consensus,
                "confidence": confidence,
                "notes": notes,
            }
        )
        if consensus == "disagree":
            artifact_root = Path(sample["artifact_root"])
            evidence = count_keyword_hits(artifact_root)
            disagreement_rows.append(
                {
                    "sample_id": sample["sample_id"],
                    "accession": accession,
                    "deeppl_label": deeppl_label,
                    "phagescope_lifestyle": phagescope_label,
                    "integrase_hits": evidence["integrase"],
                    "repressor_hits": evidence["repressor"],
                    "excisionase_hits": evidence["excisionase"],
                    "lysogeny_keyword_hits": evidence["lysogeny"],
                    "review_note": (
                        "Annotation contains lysogeny-associated keywords."
                        if sum(evidence.values()) > 0
                        else "No lysogeny-associated keywords found in saved annotations."
                    ),
                }
            )

    comparison_path = paths.integration_dir / "lifecycle_comparison.tsv"
    disagreement_path = paths.integration_dir / "disagreement_review.tsv"
    omitted_path = paths.integration_dir / "omitted_samples.tsv"
    summary_path = paths.integration_dir / "summary.json"
    write_tsv(
        comparison_path,
        ["sample_id", "accession", "phagescope_lifestyle", "deeppl_label", "consensus", "confidence", "notes"],
        comparison_rows,
    )
    write_tsv(
        disagreement_path,
        [
            "sample_id",
            "accession",
            "deeppl_label",
            "phagescope_lifestyle",
            "integrase_hits",
            "repressor_hits",
            "excisionase_hits",
            "lysogeny_keyword_hits",
            "review_note",
        ],
        disagreement_rows,
    )
    write_tsv(
        omitted_path,
        ["sample_id", "accession", "reason"],
        omitted_rows,
    )
    summary = {
        "time": now_iso(),
        "completed_save_all_samples": len(completed_samples),
        "completed_samples": len(comparison_rows),
        "agree": sum(1 for row in comparison_rows if row["consensus"] == "agree"),
        "disagree": sum(1 for row in comparison_rows if row["consensus"] == "disagree"),
        "omitted_samples": len(omitted_rows),
        "agreement_rate": round(
            sum(1 for row in comparison_rows if row["consensus"] == "agree") / max(1, len(comparison_rows)) * 100,
            4,
        ),
    }
    summary_path.write_text(json.dumps(summary, indent=2, ensure_ascii=False), encoding="utf-8")
    manifest.setdefault("paths", {}).update(
        {
            "integration_comparison": str(comparison_path.resolve()),
            "integration_disagreement_review": str(disagreement_path.resolve()),
            "integration_omitted_samples": str(omitted_path.resolve()),
            "integration_summary": str(summary_path.resolve()),
        }
    )
    mark_stage(
        manifest,
        "integration",
        "completed",
        completed=len(comparison_rows),
        disagreements=len(disagreement_rows),
        omitted=len(omitted_rows),
    )
    return comparison_rows


def apply_figure_style(fig: plt.Figure, axes: List[plt.Axes] | Tuple[plt.Axes, ...]) -> None:
    fig.patch.set_facecolor(FIGURE_BG)
    for ax in axes:
        ax.set_facecolor(PANEL_BG)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color(GRID_COLOR)
        ax.spines["bottom"].set_color(GRID_COLOR)
        ax.tick_params(colors=TEXT_PRIMARY, labelsize=11)
        ax.title.set_color(TEXT_PRIMARY)
        ax.xaxis.label.set_color(TEXT_PRIMARY)
        ax.yaxis.label.set_color(TEXT_PRIMARY)


def save_publication_figure(fig: plt.Figure, output_path: Path) -> None:
    fig.savefig(output_path, dpi=320, bbox_inches="tight", facecolor=fig.get_facecolor())
    fig.savefig(output_path.with_suffix(".svg"), bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_confusion_matrix(paths: Paths, metrics: Dict[str, Any]) -> None:
    matrix = [[metrics["tn"], metrics["fp"]], [metrics["fn"], metrics["tp"]]]
    total = max(1, metrics["n_total"])
    cmap = LinearSegmentedColormap.from_list("paper_blues", ["#edf4f8", "#a7d0e5", BLUE_DARK])
    fig, ax = plt.subplots(figsize=(7.6, 6.4))
    apply_figure_style(fig, [ax])
    image = ax.imshow(matrix, cmap=cmap, vmin=0, vmax=max(max(row) for row in matrix))
    for row_idx, row in enumerate(matrix):
        for col_idx, value in enumerate(row):
            pct = value / total * 100
            ax.text(
                col_idx,
                row_idx,
                f"{value}\n{pct:.1f}%",
                ha="center",
                va="center",
                fontsize=14,
                fontweight="bold",
                color=PANEL_BG if value > max(max(r) for r in matrix) * 0.45 else TEXT_PRIMARY,
            )
    ax.set_xticks([0, 1])
    ax.set_yticks([0, 1])
    ax.set_xticklabels(["Virulent", "Temperate"])
    ax.set_yticklabels(["Virulent", "Temperate"])
    ax.set_xlabel("Predicted label")
    ax.set_ylabel("True label")
    fig.suptitle("DeepPL confusion matrix", fontsize=20, fontweight="bold", color=TEXT_PRIMARY, y=0.98)
    fig.text(
        0.125,
        0.92,
        f"Real benchmark on {metrics['n_total']} genomes. Cell labels show count and share of the full set.",
        fontsize=11,
        color=TEXT_MUTED,
    )
    cbar = fig.colorbar(image, ax=ax, fraction=0.046, pad=0.04)
    cbar.outline.set_edgecolor(GRID_COLOR)
    cbar.ax.tick_params(labelsize=10, colors=TEXT_MUTED)
    fig.tight_layout(rect=[0, 0, 1, 0.88])
    save_publication_figure(fig, paths.figures_dir / "deeppl_confusion_matrix.png")


def plot_metrics_vs_paper(paths: Paths, metrics: Dict[str, Any]) -> None:
    metric_rows = [
        ("Accuracy", PAPER_METRICS["accuracy"], metrics["accuracy"]),
        ("Sensitivity", PAPER_METRICS["sensitivity"], metrics["sensitivity"]),
        ("Specificity", PAPER_METRICS["specificity"], metrics["specificity"]),
        ("F1", PAPER_METRICS["f1"] * 100, metrics["f1"] * 100),
        ("MCC", PAPER_METRICS["mcc"] * 100, metrics["mcc"] * 100),
    ]
    labels = [row[0] for row in metric_rows]
    paper = [row[1] for row in metric_rows]
    actual = [row[2] for row in metric_rows]
    y_positions = list(range(len(labels)))
    fig, ax = plt.subplots(figsize=(8.8, 5.8))
    apply_figure_style(fig, [ax])
    for y, paper_value, actual_value in zip(y_positions, paper, actual):
        ax.plot([paper_value, actual_value], [y, y], color=GRID_COLOR, linewidth=3, solid_capstyle="round", zorder=1)
    ax.scatter(paper, y_positions, s=120, color=SLATE, label="Paper", zorder=3)
    ax.scatter(actual, y_positions, s=140, color=BLUE_DARK, marker="D", label="Real run", zorder=4)
    for y, paper_value, actual_value in zip(y_positions, paper, actual):
        delta = actual_value - paper_value
        ax.text(actual_value + 1.4, y, f"{delta:+.2f}", va="center", fontsize=10, color=TEXT_MUTED)
    ax.set_yticks(y_positions)
    ax.set_yticklabels(labels)
    ax.set_ylim(-0.6, len(labels) - 0.4)
    ax.invert_yaxis()
    ax.set_xlim(45, 100)
    ax.set_xlabel("Score (percentage points)")
    fig.suptitle("DeepPL paper metrics vs. real benchmark", fontsize=20, fontweight="bold", color=TEXT_PRIMARY, y=0.98)
    fig.text(
        0.125,
        0.92,
        "F1 and MCC are scaled to percentage points to share the same axis with accuracy metrics.",
        fontsize=11,
        color=TEXT_MUTED,
    )
    ax.grid(axis="x", color=GRID_COLOR, alpha=0.7, linewidth=0.8)
    ax.legend(frameon=False, loc="lower right")
    fig.tight_layout(rect=[0, 0, 1, 0.88])
    save_publication_figure(fig, paths.figures_dir / "deeppl_metrics_vs_paper.png")


def plot_agreement_bar(paths: Paths, comparison_rows: List[Dict[str, Any]]) -> None:
    counts = Counter(row["consensus"] for row in comparison_rows)
    labels = ["agree", "disagree"]
    values = [counts.get(label, 0) for label in labels]
    agreement_rate = values[0] / max(1, sum(values)) * 100
    cross = Counter((row["phagescope_lifestyle"], row["deeppl_label"]) for row in comparison_rows)
    heatmap = [
        [cross.get(("virulent", "virulent"), 0), cross.get(("virulent", "temperate"), 0)],
        [cross.get(("temperate", "virulent"), 0), cross.get(("temperate", "temperate"), 0)],
    ]
    fig, (ax_left, ax_right) = plt.subplots(
        1,
        2,
        figsize=(10.8, 4.8),
        gridspec_kw={"width_ratios": [1.15, 1]},
    )
    apply_figure_style(fig, [ax_left, ax_right])
    ax_left.barh(["Agreement"], [values[0]], color=GREEN, height=0.46, label="Agree")
    ax_left.barh(["Agreement"], [values[1]], left=[values[0]], color=RED, height=0.46, label="Disagree")
    ax_left.set_xlim(0, max(1, sum(values)))
    ax_left.set_xlabel("Sample count")
    ax_left.set_title("Overall agreement", fontsize=14, fontweight="bold", pad=12)
    ax_left.text(
        values[0] / 2,
        0,
        f"{values[0]} agree",
        ha="center",
        va="center",
        color=PANEL_BG,
        fontsize=11,
        fontweight="bold",
    )
    if values[1]:
        ax_left.text(
            values[0] + values[1] / 2,
            0,
            f"{values[1]} disagree",
            ha="center",
            va="center",
            color=PANEL_BG,
            fontsize=11,
            fontweight="bold",
        )
    ax_left.text(
        0.0,
        1.08,
        f"Agreement rate: {agreement_rate:.1f}%",
        transform=ax_left.transAxes,
        fontsize=11,
        color=TEXT_MUTED,
    )
    cmap = LinearSegmentedColormap.from_list("agree_heat", ["#eef4f7", "#8ab9d6", BLUE_DARK])
    im = ax_right.imshow(heatmap, cmap=cmap, vmin=0, vmax=max(1, max(max(row) for row in heatmap)))
    for row_idx, row in enumerate(heatmap):
        for col_idx, value in enumerate(row):
            ax_right.text(
                col_idx,
                row_idx,
                str(value),
                ha="center",
                va="center",
                fontsize=13,
                fontweight="bold",
                color=PANEL_BG if value > max(1, max(max(r) for r in heatmap)) * 0.45 else TEXT_PRIMARY,
            )
    ax_right.set_xticks([0, 1])
    ax_right.set_yticks([0, 1])
    ax_right.set_xticklabels(["Virulent", "Temperate"])
    ax_right.set_yticklabels(["Virulent", "Temperate"])
    ax_right.set_xlabel("DeepPL label")
    ax_right.set_ylabel("PhageScope label")
    ax_right.set_title("Label cross-tab", fontsize=14, fontweight="bold", pad=12)
    cbar = fig.colorbar(im, ax=ax_right, fraction=0.046, pad=0.04)
    cbar.outline.set_edgecolor(GRID_COLOR)
    cbar.ax.tick_params(labelsize=9, colors=TEXT_MUTED)
    fig.suptitle("PhageScope vs. DeepPL lifecycle agreement", fontsize=20, fontweight="bold", color=TEXT_PRIMARY, y=0.99)
    fig.tight_layout(rect=[0, 0, 1, 0.94])
    save_publication_figure(fig, paths.figures_dir / "phagescope_deeppl_agreement_bar.png")


def plot_disagreement_evidence(paths: Paths, disagreement_rows: List[Dict[str, Any]]) -> None:
    if not disagreement_rows:
        fig, ax = plt.subplots(figsize=(8.5, 3.8))
        apply_figure_style(fig, [ax])
        ax.axis("off")
        ax.text(0.5, 0.62, "No disagreement cases required evidence review.", ha="center", va="center", fontsize=18, fontweight="bold", color=TEXT_PRIMARY)
        ax.text(0.5, 0.38, "The integration subset showed full agreement across all comparable samples.", ha="center", va="center", fontsize=11, color=TEXT_MUTED)
        save_publication_figure(fig, paths.figures_dir / "disagreement_evidence_overview.png")
        return

    totals = {key: 0 for key in ("integrase_hits", "repressor_hits", "excisionase_hits", "lysogeny_keyword_hits")}
    per_sample_totals = []
    for row in disagreement_rows:
        total_hits = 0
        for key in totals:
            count = int(row.get(key) or 0)
            totals[key] += count
            total_hits += count
        per_sample_totals.append((row["sample_id"], total_hits))
    per_sample_totals.sort(key=lambda item: (-item[1], item[0]))
    if sum(totals.values()) == 0:
        fig, ax = plt.subplots(figsize=(9.2, 4.0))
        apply_figure_style(fig, [ax])
        ax.axis("off")
        ax.text(0.5, 0.64, "Disagreement cases lacked lysogeny-associated keywords.", ha="center", va="center", fontsize=18, fontweight="bold", color=TEXT_PRIMARY)
        ax.text(
            0.5,
            0.39,
            f"{len(disagreement_rows)} disagreement cases were reviewed, but integrase/repressor/excisionase and broader lysogeny keywords were absent in saved annotations.",
            ha="center",
            va="center",
            fontsize=11,
            color=TEXT_MUTED,
            wrap=True,
        )
        save_publication_figure(fig, paths.figures_dir / "disagreement_evidence_overview.png")
        return
    fig, (ax_left, ax_right) = plt.subplots(
        1,
        2,
        figsize=(11.2, 4.9),
        gridspec_kw={"width_ratios": [1, 1.15]},
    )
    apply_figure_style(fig, [ax_left, ax_right])
    labels = ["Integrase", "Repressor", "Excisionase", "Lysogeny keywords"]
    values = [totals["integrase_hits"], totals["repressor_hits"], totals["excisionase_hits"], totals["lysogeny_keyword_hits"]]
    colors = [BLUE_DARK, BLUE_LIGHT, GOLD, RED]
    ax_left.barh(labels, values, color=colors, height=0.56)
    ax_left.invert_yaxis()
    ax_left.set_xlabel("Keyword hits")
    ax_left.set_title("Evidence categories", fontsize=14, fontweight="bold", pad=12)
    for idx, value in enumerate(values):
        ax_left.text(value + 0.15, idx, str(value), va="center", fontsize=10, color=TEXT_MUTED)
    sample_labels = [item[0] for item in per_sample_totals]
    sample_values = [item[1] for item in per_sample_totals]
    ax_right.barh(sample_labels, sample_values, color=GOLD, height=0.56)
    ax_right.invert_yaxis()
    ax_right.set_xlabel("Total evidence hits")
    ax_right.set_title("Per-sample evidence burden", fontsize=14, fontweight="bold", pad=12)
    for idx, value in enumerate(sample_values):
        ax_right.text(value + 0.15, idx, str(value), va="center", fontsize=10, color=TEXT_MUTED)
    fig.suptitle("Disagreement evidence overview", fontsize=20, fontweight="bold", color=TEXT_PRIMARY, y=0.99)
    fig.tight_layout(rect=[0, 0, 1, 0.94])
    save_publication_figure(fig, paths.figures_dir / "disagreement_evidence_overview.png")


def build_figures(paths: Paths, metrics: Dict[str, Any], comparison_rows: List[Dict[str, Any]], disagreement_rows: List[Dict[str, Any]]) -> None:
    plot_confusion_matrix(paths, metrics)
    plot_metrics_vs_paper(paths, metrics)
    plot_agreement_bar(paths, comparison_rows)
    plot_disagreement_evidence(paths, disagreement_rows)


def write_methods_runbook(paths: Paths, manifest: Dict[str, Any], metrics: Dict[str, Any], comparison_rows: List[Dict[str, Any]], completed_samples: List[Dict[str, Any]]) -> None:
    provenance = {}
    provenance_path = paths.deeppl_dir / "run_provenance.json"
    if provenance_path.exists():
        provenance = json.loads(provenance_path.read_text(encoding="utf-8"))
    text = textwrap.dedent(
        f"""
        # Methods Runbook

        - Generated at: {now_iso()}
        - Output root: {paths.root}
        - Ground truth sources:
          - {TEST_LABELS_CSV}
          - {SUPPLEMENTAL_XLSX}
        - Ground truth rows: {manifest.get('counts', {}).get('ground_truth_rows', 0)}
        - DeepPL benchmark rows: {manifest.get('counts', {}).get('deeppl_predictions', 0)}
        - DeepPL skipped source rows: {manifest.get('counts', {}).get('deeppl_skipped', 0)}
        - PhageScope completed validation samples: {len(completed_samples)}
        - Integration comparison rows: {len(comparison_rows)}

        ## DeepPL benchmark

        - Download method: NCBI EFetch ({EFETCH_URL})
        - Model path: {provenance.get('remote_model_path') or DEFAULT_DEEPPL_MODEL_DIR}
        - Execution resource: remote profile `{provenance.get('remote_profile', 'gpu')}`
        - Remote host: {provenance.get('remote_host', 'unknown')}
        - Remote project: {provenance.get('remote_project', 'unknown')}
        - Remote script: {provenance.get('remote_script', 'unknown')}
        - Output CSV: {paths.deeppl_raw_dir / 'deeppl_batch_predictions.csv'}

        ## PhageScope validation

        - Submit mode: FASTA upload
        - Analysis type: Annotation Pipline
        - Modules: quality, host, lifestyle, annotation
        - Completed samples: {len(completed_samples)}
        - Save-all root: {paths.phagescope_saveall_dir}

        ## Label normalization

        - Lysogenic / Temperate / Lysogen* -> temperate
        - Lytic / Virulent -> virulent
        - N bases are removed before DeepPL input generation
        - Non-ATCG IUPAC ambiguity codes are deterministically replaced when count <= 10 per genome
        - Source rows that still cannot be retrieved as nucleotide FASTA are recorded in `data/download_skipped.tsv`

        ## Real benchmark metrics

        - Accuracy: {metrics['accuracy']}
        - Sensitivity: {metrics['sensitivity']}
        - Specificity: {metrics['specificity']}
        - Precision: {metrics['precision']}
        - F1: {metrics['f1']}
        - MCC: {metrics['mcc']}
        """
    ).strip() + "\n"
    (paths.methods_dir / "methods_runbook.md").write_text(text, encoding="utf-8")


def write_results_summary(paths: Paths, metrics: Dict[str, Any], comparison_rows: List[Dict[str, Any]], disagreement_rows: List[Dict[str, Any]]) -> None:
    agree = sum(1 for row in comparison_rows if row["consensus"] == "agree")
    disagree = sum(1 for row in comparison_rows if row["consensus"] == "disagree")
    integration_summary = {}
    summary_path = paths.integration_dir / "summary.json"
    if summary_path.exists():
        integration_summary = json.loads(summary_path.read_text(encoding="utf-8"))
    text = textwrap.dedent(
        f"""
        # Results Summary

        ## DeepPL real benchmark
        - Samples: {metrics['n_total']}
        - Skipped source rows: {json.loads((paths.root / 'manifest.json').read_text(encoding='utf-8')).get('counts', {}).get('deeppl_skipped', 0)}
        - Accuracy: {metrics['accuracy']} (paper {PAPER_METRICS['accuracy']})
        - Sensitivity: {metrics['sensitivity']} (paper {PAPER_METRICS['sensitivity']})
        - Specificity: {metrics['specificity']} (paper {PAPER_METRICS['specificity']})
        - F1: {metrics['f1']}
        - MCC: {metrics['mcc']} (paper {PAPER_METRICS['mcc']})

        ## PhageScope integration subset
        - Save-all completed samples: {integration_summary.get('completed_save_all_samples', len(comparison_rows))}
        - Comparable samples with lifecycle labels: {len(comparison_rows)}
        - Omitted samples: {integration_summary.get('omitted_samples', 0)}
        - Agree: {agree}
        - Disagree: {disagree}
        - Agreement rate: {round(agree / max(1, len(comparison_rows)) * 100, 4)}

        ## Disagreement review
        - Cases requiring review: {len(disagreement_rows)}
        - Evidence file: {paths.integration_dir / 'disagreement_review.tsv'}
        """
    ).strip() + "\n"
    (paths.methods_dir / "results_summary.md").write_text(text, encoding="utf-8")


def run_all(args: argparse.Namespace) -> int:
    output_root = Path(args.output_root).expanduser().resolve()
    paths = build_paths(output_root)
    init_layout(paths)
    manifest = load_manifest(paths.manifest_path)
    save_manifest(paths.manifest_path, manifest)

    try:
        truth_rows = build_ground_truth(paths, manifest)
        save_manifest(paths.manifest_path, manifest)

        download_rows = download_test_set(paths, manifest, truth_rows)
        save_manifest(paths.manifest_path, manifest)

        benchmark_rows = build_deeppl_benchmark(
            paths,
            manifest,
            truth_rows,
            download_rows,
            remote_profile=args.deeppl_remote_profile,
            timeout_sec=args.deeppl_timeout,
        )
        save_manifest(paths.manifest_path, manifest)

        subset_rows = build_subset_manifest(paths, benchmark_rows, args.phagescope_per_class)
        save_manifest(paths.manifest_path, manifest)

        completed_samples, _registry = asyncio.run(
            run_phagescope_validation(
                paths,
                manifest,
                subset_rows,
                poll_interval=args.phagescope_poll_interval,
                poll_timeout=args.phagescope_poll_timeout,
            )
        )
        save_manifest(paths.manifest_path, manifest)

        comparison_rows = build_integration_outputs(paths, manifest, benchmark_rows, completed_samples)
        disagreement_rows = read_tsv(paths.integration_dir / "disagreement_review.tsv")
        metrics = json.loads((paths.deeppl_dir / "benchmark_metrics.json").read_text(encoding="utf-8"))
        build_figures(paths, metrics, comparison_rows, disagreement_rows)
        write_methods_runbook(paths, manifest, metrics, comparison_rows, completed_samples)
        write_results_summary(paths, metrics, comparison_rows, disagreement_rows)
        manifest.setdefault("paths", {}).update(
            {
                "methods_runbook": str((paths.methods_dir / "methods_runbook.md").resolve()),
                "results_summary": str((paths.methods_dir / "results_summary.md").resolve()),
                "figure_confusion_matrix": str((paths.figures_dir / "deeppl_confusion_matrix.png").resolve()),
                "figure_metrics_vs_paper": str((paths.figures_dir / "deeppl_metrics_vs_paper.png").resolve()),
                "figure_agreement_bar": str((paths.figures_dir / "phagescope_deeppl_agreement_bar.png").resolve()),
                "figure_disagreement_evidence": str((paths.figures_dir / "disagreement_evidence_overview.png").resolve()),
            }
        )
        mark_stage(manifest, "package", "completed")
        save_manifest(paths.manifest_path, manifest)
        print(f"Paper asset package ready at {paths.root}")
        return 0
    except Exception as exc:
        append_failure(manifest, "pipeline", str(exc))
        mark_stage(manifest, "package", "failed", error=str(exc))
        save_manifest(paths.manifest_path, manifest)
        raise


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build a real-run DeepPL × PhageScope paper asset package.")
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT), help="Output package root directory.")
    parser.add_argument("--deeppl-remote-profile", default="gpu", choices=["gpu", "cpu", "default"], help="Remote DeepPL profile.")
    parser.add_argument("--deeppl-timeout", type=int, default=21600, help="DeepPL remote batch timeout in seconds.")
    parser.add_argument("--phagescope-per-class", type=int, default=24, help="Validation subset size per lifecycle class.")
    parser.add_argument("--phagescope-poll-interval", type=float, default=30.0, help="Polling interval in seconds for PhageScope task_detail.")
    parser.add_argument("--phagescope-poll-timeout", type=float, default=1800.0, help="Total polling timeout in seconds for PhageScope validation.")
    return parser


def main() -> int:
    parser = build_arg_parser()
    args = parser.parse_args()
    return run_all(args)


if __name__ == "__main__":
    raise SystemExit(main())
